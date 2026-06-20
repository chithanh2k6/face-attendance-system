import csv
import os
from datetime import datetime

from modules.database import (
    create_tables,
    get_all_attendance,
    get_attendance_by_date,
    get_attendance_by_date_range,
    REPORTS_DIR,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


# ──────────────────────────────────────────────
# Cấu hình báo cáo
# ──────────────────────────────────────────────

CSV_HEADERS = [
    "MSSV",
    "Họ tên",
    "Lớp",
    "Ngày",
    "Giờ",
    "Trạng thái",
    "Môn học",
    "Ghi chú",
]



# ──────────────────────────────────────────────
# Xử lý tên file báo cáo
# ──────────────────────────────────────────────

def normalize_file_format(file_format):
    """
    Chuẩn hóa định dạng file báo cáo.
    Hỗ trợ: csv, excel, xlsx.
    """
    file_format = file_format.strip().lower() if file_format else "csv"

    if file_format == "xlsx":
        file_format = "excel"

    if file_format not in ["csv", "excel"]:
        file_format = "csv"

    return file_format


def get_file_extension(file_format):
    """
    Lấy phần mở rộng file theo định dạng báo cáo.
    """
    file_format = normalize_file_format(file_format)
    return "xlsx" if file_format == "excel" else "csv"


def generate_report_filename(prefix="attendance_report", extension="csv"):
    """
    Tạo tên file báo cáo tự động theo thời gian hiện tại.
    Tránh ghi đè file báo cáo cũ.

    Ví dụ:
        attendance_report_2026-05-27_143022.csv
        attendance_report_2026-05-27_143022.xlsx
    """
    extension = extension.strip().lstrip(".") if extension else "csv"
    current_time = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"{prefix}_{current_time}.{extension}"
    return os.path.join(REPORTS_DIR, filename)


def ensure_reports_dir():
    """
    Đảm bảo thư mục reports/ tồn tại trước khi xuất file.
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)


# ──────────────────────────────────────────────
# Kiểm tra ngày tháng
# ──────────────────────────────────────────────

def is_valid_date(date):
    """
    Kiểm tra ngày có đúng định dạng YYYY-MM-DD hay không.
    """
    try:
        datetime.strptime(date, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def validate_date_range(start_date, end_date):
    """
    Kiểm tra khoảng ngày hợp lệ.
    Trả về: (True, "") nếu hợp lệ, ngược lại trả về (False, message).
    """
    start_date = start_date.strip() if start_date else ""
    end_date   = end_date.strip() if end_date else ""

    if not start_date or not end_date:
        return False, "Vui lòng nhập đầy đủ ngày bắt đầu và ngày kết thúc."

    if not is_valid_date(start_date) or not is_valid_date(end_date):
        return False, "Ngày phải đúng định dạng YYYY-MM-DD."

    if start_date > end_date:
        return False, "Ngày bắt đầu không được lớn hơn ngày kết thúc."

    return True, ""


# ──────────────────────────────────────────────
# Chuẩn hóa dữ liệu báo cáo
# ──────────────────────────────────────────────

def record_to_row(row):
    """
    Chuyển 1 dòng dữ liệu điểm danh dạng dict thành list để ghi file.
    """
    return [
        row.get("student_id", ""),
        row.get("full_name", ""),
        row.get("class_name", ""),
        row.get("date", ""),
        row.get("time", ""),
        row.get("status", ""),
        row.get("subject", ""),
        row.get("note", ""),
    ]


# ──────────────────────────────────────────────
# Xuất dữ liệu ra CSV
# ──────────────────────────────────────────────

def export_records_to_csv(records, output_file):
    """
    Ghi danh sách điểm danh ra file CSV.

    Tham số:
        records: list dict dữ liệu điểm danh.
        output_file: đường dẫn file CSV cần xuất.

    Trả về:
        (True, message) nếu xuất thành công.
        (False, message) nếu không có dữ liệu hoặc lỗi ghi file.
    """
    if not records:
        return False, "Không có dữ liệu điểm danh để xuất báo cáo."

    ensure_reports_dir()

    try:
        with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(CSV_HEADERS)

            for row in records:
                writer.writerow(record_to_row(row))

        return True, f"Xuất báo cáo CSV thành công: {output_file}"

    except Exception as e:
        return False, f"Lỗi khi xuất báo cáo CSV: {e}"


# ──────────────────────────────────────────────
# Xuất dữ liệu ra Excel
# ──────────────────────────────────────────────

def export_records_to_excel(records, output_file):
    """
    Ghi danh sách điểm danh ra file Excel (.xlsx).

    Trả về:
        (True, message) nếu xuất thành công.
        (False, message) nếu không có dữ liệu hoặc lỗi ghi file.
    """
    if not records:
        return False, "Không có dữ liệu điểm danh để xuất báo cáo."

    if Workbook is None:
        return False, "Không thể xuất file Excel."

    ensure_reports_dir()

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "DiemDanh"

        ws.append(CSV_HEADERS)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in records:
            ws.append(record_to_row(row))

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in column_cells:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 35)

        ws.freeze_panes = "A2"
        wb.save(output_file)

        return True, f"Xuất báo cáo Excel thành công: {output_file}"

    except Exception as e:
        return False, f"Lỗi khi xuất báo cáo Excel: {e}"


# ──────────────────────────────────────────────
# Điều phối xuất báo cáo
# ──────────────────────────────────────────────

def export_records(records, output_file, file_format="csv"):
    """
    Xuất danh sách điểm danh theo định dạng CSV hoặc Excel.
    """
    file_format = normalize_file_format(file_format)

    if file_format == "excel":
        return export_records_to_excel(records, output_file)

    return export_records_to_csv(records, output_file)


def export_all_attendance(output_file=None, file_format="csv"):
    """
    Xuất toàn bộ lịch sử điểm danh ra CSV hoặc Excel.
    """
    create_tables()
    file_format = normalize_file_format(file_format)

    if output_file is None:
        extension   = get_file_extension(file_format)
        output_file = generate_report_filename("attendance_all", extension)

    records = get_all_attendance()
    return export_records(records, output_file, file_format)


def export_attendance_by_date(date, output_file=None, file_format="csv"):
    """
    Xuất danh sách điểm danh theo ngày ra CSV hoặc Excel.

    Tham số:
        date: ngày cần xuất theo định dạng YYYY-MM-DD.
    """
    create_tables()
    file_format = normalize_file_format(file_format)

    date = date.strip() if date else ""

    if not date or not is_valid_date(date):
        return False, "Ngày xuất báo cáo không hợp lệ. Vui lòng nhập dạng YYYY-MM-DD."

    if output_file is None:
        extension   = get_file_extension(file_format)
        output_file = generate_report_filename(f"attendance_{date}", extension)

    records = get_attendance_by_date(date)
    return export_records(records, output_file, file_format)


def export_attendance_by_date_range(start_date, end_date, output_file=None, file_format="csv"):
    """
    Xuất danh sách điểm danh theo khoảng ngày ra CSV hoặc Excel.

    Tham số:
        start_date: ngày bắt đầu theo định dạng YYYY-MM-DD.
        end_date: ngày kết thúc theo định dạng YYYY-MM-DD.
    """
    create_tables()
    file_format = normalize_file_format(file_format)

    valid, message = validate_date_range(start_date, end_date)
    if not valid:
        return False, message

    start_date = start_date.strip()
    end_date   = end_date.strip()

    if output_file is None:
        extension   = get_file_extension(file_format)
        output_file = generate_report_filename(f"attendance_{start_date}_to_{end_date}", extension)

    records = get_attendance_by_date_range(start_date, end_date)
    return export_records(records, output_file, file_format)


def export_attendance_report(report_type="all", file_format="csv",
                             date="", start_date="", end_date=""):
    """
    Hàm tổng hợp để UI gọi xuất báo cáo.

    report_type:
        all   : xuất toàn bộ
        date  : xuất theo ngày
        range : xuất theo khoảng ngày

    file_format:
        csv hoặc excel
    """
    report_type = report_type.strip().lower() if report_type else "all"

    if report_type == "date":
        return export_attendance_by_date(date, file_format=file_format)

    if report_type == "range":
        return export_attendance_by_date_range(start_date, end_date, file_format=file_format)

    return export_all_attendance(file_format=file_format)


# ──────────────────────────────────────────────
# Hàm in dữ liệu ra terminal để test nhanh
# ──────────────────────────────────────────────

def print_report_preview(records):
    """
    In thử dữ liệu báo cáo ra terminal.
    Dùng để kiểm tra trước khi export CSV.
    """
    if not records:
        print("Không có dữ liệu để hiển thị.")
        return

    print("\n=== PREVIEW BÁO CÁO ĐIỂM DANH ===")
    print(f"{'MSSV':<15} {'Họ tên':<25} {'Lớp':<15} {'Ngày':<12} {'Giờ':<10} {'Trạng thái':<12}")

    for row in records:
        print(
            f"{row.get('student_id', ''):<15} "
            f"{row.get('full_name', ''):<25} "
            f"{row.get('class_name', '') or '':<15} "
            f"{row.get('date', ''):<12} "
            f"{row.get('time', ''):<10} "
            f"{row.get('status', ''):<12}"
        )


# ──────────────────────────────────────────────
# Flow test CLI
# ──────────────────────────────────────────────

def report_cli():
    """
    Menu test nhanh report.py bằng terminal.
    Chạy bằng lệnh:
        python -m modules.report
    """
    create_tables()

    while True:
        print("\n=== TEST REPORT ===")
        print("1. Xem preview toàn bộ điểm danh")
        print("2. Xuất toàn bộ điểm danh ra CSV")
        print("3. Xuất điểm danh theo ngày ra CSV")
        print("4. Xuất điểm danh theo khoảng ngày ra CSV")
        print("5. Xuất toàn bộ điểm danh ra Excel")
        print("6. Xuất điểm danh theo ngày ra Excel")
        print("7. Xuất điểm danh theo khoảng ngày ra Excel")
        print("0. Thoát")

        choice = input("Chọn chức năng: ").strip()

        if choice == "1":
            records = get_all_attendance()
            print_report_preview(records)

        elif choice == "2":
            success, message = export_all_attendance(file_format="csv")
            print(message)

        elif choice == "3":
            date = input("Nhập ngày cần xuất (YYYY-MM-DD): ").strip()
            success, message = export_attendance_by_date(date, file_format="csv")
            print(message)

        elif choice == "4":
            start_date = input("Nhập ngày bắt đầu (YYYY-MM-DD): ").strip()
            end_date   = input("Nhập ngày kết thúc (YYYY-MM-DD): ").strip()
            success, message = export_attendance_by_date_range(start_date, end_date, file_format="csv")
            print(message)

        elif choice == "5":
            success, message = export_all_attendance(file_format="excel")
            print(message)

        elif choice == "6":
            date = input("Nhập ngày cần xuất (YYYY-MM-DD): ").strip()
            success, message = export_attendance_by_date(date, file_format="excel")
            print(message)

        elif choice == "7":
            start_date = input("Nhập ngày bắt đầu (YYYY-MM-DD): ").strip()
            end_date   = input("Nhập ngày kết thúc (YYYY-MM-DD): ").strip()
            success, message = export_attendance_by_date_range(start_date, end_date, file_format="excel")
            print(message)

        elif choice == "0":
            print("Đã thoát.")
            break

        else:
            print("Lựa chọn không hợp lệ.")


# ──────────────────────────────────────────────
# Test khi chạy trực tiếp
# ──────────────────────────────────────────────

if __name__ == "__main__":
    report_cli()
